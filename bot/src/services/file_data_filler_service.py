"""
FileDataFillerService - LLM phân tích file Excel + sinh SQL + fill data
"""

import os
import json
import re
from datetime import datetime
from typing import Any
import tempfile

from src.services.excel_processor_service import get_excel_processor, ExcelProcessorService
from src.services.sql_assistant_service import get_sql_assistant_service
from src.services.llm_service import get_llm_service
from src.db import db
from src.utilities import get_logger

logger = get_logger(__name__)


# Prompt template cho LLM phân tích file và sinh SQL
FILE_ANALYSIS_PROMPT = """
Bạn là chuyên gia phân tích dữ liệu và thiết kế truy vấn SQL cho file Excel.

Nhiệm vụ:
1. Đọc cấu trúc file Excel (headers/cột).
2. Xác định ý nghĩa từng cột.
3. Sinh SQL để lấy dữ liệu tương ứng.
4. Trả về mapping cột Excel -> cột dữ liệu.

==================================================
CẤU TRÚC FILE EXCEL
==================================================

{file_structure}

==================================================
DATABASE SCHEMA
==================================================

{schema}

==================================================
RÀNG BUỘC BẮT BUỘC
==================================================

1. CHỈ được dùng 3 bảng:
   - "ProjectRegistration"
   - "User"
   - "CallRound"

2. SQL sinh ra phải:
   - Chỉ SELECT (không INSERT/UPDATE/DELETE/ALTER/DROP).
   - Chỉ dùng bảng/cột tồn tại trong schema.
   - Dùng JOIN đúng quan hệ.

3. Mapping cột:
   - `columnMappings` có key là TẤT CẢ cột Excel.
   - Value là tên cột trả về trong SQL (alias) hoặc null.
   - Cột nào không map được bắt buộc trả về null.

4. Trả về JSON hợp lệ theo format:
{{
  "analysis": "Mô tả ngắn",
  "dataType": "PROJECT_REGISTRATION | USER | CALL_ROUND | OTHER",
  "sql": "SELECT ... FROM ...",
  "columnMappings": {{
    "<ten_cot_excel_1>": "<ten_alias_sql_1>",
    "<ten_cot_excel_2>": null
  }},
  "filters": ["mô tả filter nếu có"]
}}

==================================================
VÍ DỤ
==================================================

File có cột:
- Mã SV
- Tên đề tài
- Trạng thái
- Ghi chú

Trả về:
{{
  "analysis": "Danh sách sinh viên đăng ký đề tài",
  "dataType": "PROJECT_REGISTRATION",
  "sql": "SELECT pr.\"title\" AS \"Tên đề tài\", u.\"name\" AS \"Tên sinh viên\", u.\"code\" AS \"Mã SV\", pr.\"status\" AS \"Trạng thái\" FROM \"ProjectRegistration\" pr JOIN \"User\" u ON pr.\"userId\" = u.\"id\"",
  "columnMappings": {{
    "Mã SV": "Mã SV",
    "Tên đề tài": "Tên đề tài",
    "Trạng thái": "Trạng thái",
    "Ghi chú": null
  }},
  "filters": []
}}
"""

HEADER_REMAP_PROMPT = """
Bạn là chuyên gia mapping cột Excel.

Input:
- excelHeaders: danh sách header thật trong file
- currentMappings: mapping hiện tại (có thể sai key)
- dataKeys: danh sách key thực tế trả về từ SQL

Yêu cầu:
1) Map lại để key mapping KHỚP CHÍNH XÁC với excelHeaders.
2) Value phải là 1 key trong dataKeys hoặc null.
3) Nếu currentMappings có key dạng Sheet.Column_x hoặc typo gần giống header, hãy suy luận map sang header đúng.
4) Trả về JSON:
{
  "columnMappings": {
    "<excelHeader>": "<dataKey|null>"
  }
}
"""


class FileDataFillerService:
    """
    Service de fill du lieu vao file Excel su dung LLM.
    
    Flow:
    1. Doc va phan tich cau truc file Excel
    2. Su dung LLM de xac dinh y nghia cac cot
    3. Sinh SQL de lay du lieu tu database
    4. Fill du lieu vao file va tra ve
    """

    def __init__(self):
        self.excel_processor = get_excel_processor()
        self.sql_service = get_sql_assistant_service()

    async def analyze_and_fill(self, file_path: str, output_dir: str = None, call_round_id: str | None = None) -> dict[str, Any]:
        """
        Phan tich file va fill du lieu.
        
        Args:
            file_path: Duong dan file Excel can xu ly
            output_dir: Thu muc chua file output (mac dinh: /tmp)
            
        Returns:
            {
                "success": bool,
                "outputPath": str,
                "analysis": {...},
                "rowCount": int,
                "error": str (neu co)
            }
        """
        logger.info(f"🚀 Starting file data fill: {file_path}")
        
        try:
            # 1. Validate file
            if not self.excel_processor.is_valid_excel(file_path):
                raise ValueError(f"File khong phai Excel: {file_path}")
            
            # 2. Phan tich cau truc file
            analysis = self.excel_processor.analyze_excel(file_path)
            file_description = self.excel_processor.describe_for_llm(analysis)
            logger.info(f"📊 File analysis:\n{file_description}")
            
            # 3. Su dung LLM de phan tich va sinh SQL
            llm_result = await self._analyze_with_llm(file_description, call_round_id=call_round_id)
            logger.info(f"🧠 LLM analysis: {llm_result.get('dataType')}")
            
            # 4. Execute SQL
            sql = llm_result.get("sql", "")
            if sql:
                logger.info(f"🧾 Generated SQL:\n{sql}")
                logger.info(f"🧩 Column mappings: {json.dumps(llm_result.get('columnMappings', {}), ensure_ascii=False)}")
                data = await self._execute_query(sql)
                logger.info(f"📦 Fetched {len(data)} rows from database")
                if data:
                    logger.info(f"🔎 First row data: {json.dumps(data[0], ensure_ascii=False, default=str)}")
                    logger.info(f"🔎 Data keys: {list(data[0].keys())}")
                else:
                    logger.warning("⚠️ Query returned 0 rows")
            else:
                data = []

            # 4.1 Remap columnMappings theo header thật + data keys
            try:
                excel_headers = self._extract_excel_headers(file_path)
                data_keys = list(data[0].keys()) if data else []
                if excel_headers and llm_result.get("columnMappings"):
                    remapped = await self._remap_mappings_with_llm(
                        excel_headers=excel_headers,
                        current_mappings=llm_result.get("columnMappings", {}),
                        data_keys=data_keys,
                    )
                    if remapped:
                        llm_result["columnMappings"] = remapped
                        logger.info(f"🧭 Remapped mappings: {json.dumps(remapped, ensure_ascii=False)}")
            except Exception as remap_err:
                logger.warning(f"⚠️ Remap mappings skipped: {remap_err}")
            
            # 5. Fill data vao file
            output_path = self._fill_excel(file_path, data, llm_result, output_dir)
            
            return {
                "success": True,
                "outputPath": output_path,
                "analysis": {
                    "fileStructure": analysis,
                    "llmAnalysis": llm_result,
                    "dataType": llm_result.get("dataType"),
                    "sql": sql
                },
                "rowCount": len(data),
                "message": f"Da fill {len(data)} dong du lieu"
            }
            
        except Exception as e:
            logger.error(f"❌ FileDataFiller error: {e}")
            return {
                "success": False,
                "error": str(e),
                "outputPath": None,
                "analysis": None,
                "rowCount": 0
            }

    async def _analyze_with_llm(self, file_description: str, call_round_id: str | None = None) -> dict[str, Any]:
        """Su dung LLM de phan tich file va sinh SQL."""
        llm = get_llm_service()
        schema = self.sql_service._build_schema_for_prompt()
        
        prompt = FILE_ANALYSIS_PROMPT.format(
            file_structure=file_description,
            schema=schema
        )
        
        if call_round_id:
            prompt += (
                "\n\nRÀNG BUỘC FILTER BẮT BUỘC:\n"
                f"- Chỉ lấy dữ liệu thuộc callRoundId = '{call_round_id}'.\n"
                "- Nếu query có bảng ProjectRegistration/Project/CallRound thì bắt buộc WHERE theo callRoundId tương ứng.\n"
            )

        response = await llm.chat_completion(
            messages=[
                {"role": "system", "content": "You are a data analyst. Always respond with valid JSON only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            response_format={"type": "json_object"}
        )
        
        # Parse JSON response
        response_clean = response.strip()
        try:
            result = json.loads(response_clean)
        except json.JSONDecodeError:
            logger.error(f"LLM returned invalid JSON: {response_clean[:200]}")
            result = {
                "analysis": "Khong the phan tich file",
                "dataType": "UNKNOWN",
                "sql": "",
                "columnMappings": {},
                "filters": []
            }
        
        result["sql"] = self._enforce_call_round_filter(result.get("sql", ""), call_round_id)
        return result

    @staticmethod
    def _enforce_call_round_filter(sql: str, call_round_id: str | None) -> str:
        if not sql or not call_round_id:
            return sql
        lowered = sql.lower()
        if "callroundid" in lowered and call_round_id in sql:
            return sql
        candidate_aliases = ["pr", "p", "c"]
        alias = next((a for a in candidate_aliases if f'{a}."callRoundId"' in sql), "pr")
        condition = f'{alias}."callRoundId" = \'{call_round_id}\''
        if re.search(r"\bwhere\b", lowered):
            return re.sub(r"\bwhere\b", f"WHERE {condition} AND ", sql, count=1, flags=re.IGNORECASE)
        return f"{sql.rstrip().rstrip(';')} WHERE {condition};"

    async def _execute_query(self, sql: str) -> list[dict[str, Any]]:
        """Execute SQL query and return results."""
        if not sql:
            return []
        
        try:
            await db.connect()
            rows = await db.fetch(sql)
            return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"❌ SQL execution error: {e}")
            return []

    def _extract_excel_headers(self, template_path: str) -> list[str]:
        from openpyxl import load_workbook

        wb = load_workbook(template_path)
        ws = wb.active
        best_row = 1
        best_score = -1
        for r in range(1, min(ws.max_row, 60) + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            texts = [str(v).strip() for v in vals if isinstance(v, str) and str(v).strip()]
            score = len(texts)
            if any(t in {"TT", "STT", "Tên đề tài NCKH", "Tên Cán bộ/GV", "Mã đối tượng", "Mã đội tượng"} for t in texts):
                score += 5
            if score > best_score:
                best_score = score
                best_row = r

        headers = [ws.cell(row=best_row, column=c).value for c in range(1, ws.max_column + 1)]
        wb.close()
        return [str(h).strip() for h in headers if h is not None and str(h).strip()]

    async def _remap_mappings_with_llm(
        self,
        excel_headers: list[str],
        current_mappings: dict[str, Any],
        data_keys: list[str],
    ) -> dict[str, Any]:
        llm = get_llm_service()
        payload = {
            "excelHeaders": excel_headers,
            "currentMappings": current_mappings,
            "dataKeys": data_keys,
        }
        response = await llm.chat_completion(
            messages=[
                {"role": "system", "content": "Always respond valid JSON only."},
                {"role": "user", "content": HEADER_REMAP_PROMPT + "\n\n" + json.dumps(payload, ensure_ascii=False)},
            ],
            temperature=0,
            response_format={"type": "json_object"},
        )
        try:
            parsed = json.loads(response.strip())
            mapped = parsed.get("columnMappings", {}) if isinstance(parsed, dict) else {}
        except Exception:
            mapped = {}

        if not mapped:
            # fallback local: giữ key nào khớp header
            normalized = {}
            for k, v in (current_mappings or {}).items():
                key = str(k).strip()
                if "." in key:
                    key = key.split(".", 1)[1].strip()
                if key in excel_headers:
                    normalized[key] = v
            return normalized

        # sanitize output
        safe = {}
        data_key_set = set(data_keys)
        for h in excel_headers:
            v = mapped.get(h)
            safe[h] = v if (v in data_key_set) else None
        return safe

    def _fill_excel(
        self, 
        template_path: str, 
        data: list[dict[str, Any]], 
        llm_result: dict[str, Any],
        output_dir: str = None
    ) -> str:
        """Fill data vao Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            logger.warning("openpyxl not available, saving as JSON")
            return self._save_as_json(template_path, data, output_dir)
        
        wb = load_workbook(template_path)
        ws = wb.active
        
        column_mappings_raw = llm_result.get("columnMappings", {})

        # Normalize mapping keys: "Sheet.Col" -> "Col"
        column_mappings: dict[str, Any] = {}
        for k, v in column_mappings_raw.items():
            key = str(k).strip()
            if "." in key:
                key = key.split(".", 1)[1].strip()
            column_mappings[key] = v
        
        # Detect header row (avoid hardcode row=1)
        mapping_keys = {str(k).strip() for k in column_mappings.keys() if k is not None and str(k).strip()}
        header_row = 1
        best_score = -1
        for r in range(1, min(ws.max_row, 30) + 1):
            row_headers = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            normalized = {str(h).strip() for h in row_headers if h is not None and str(h).strip()}
            # Ưu tiên row có nhiều key trùng mapping + có cấu trúc header thật (>=3 ô có text)
            text_cells = [h for h in row_headers if isinstance(h, str) and h.strip()]
            score = len(normalized & mapping_keys) if mapping_keys else len(normalized)
            if len(text_cells) >= 3:
                score += 1
            if score > best_score:
                best_score = score
                header_row = r

        # Fallback cho template có nhiều dòng tiêu đề (trường hợp best_score thấp -> dính row logo/tên trường)
        if best_score <= 1:
            strong_markers = {
                "TT", "STT", "Mã đối tượng", "Họ và tên", "Tên Cán bộ/GV",
                "Tên đề tài", "Trạng thái", "Ghi chú"
            }
            for r in range(1, min(ws.max_row, 60) + 1):
                row_headers = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                normalized = {str(h).strip() for h in row_headers if h is not None and str(h).strip()}
                if len(normalized & strong_markers) >= 2:
                    header_row = r
                    logger.info(f"Header fallback matched at row={header_row}")
                    break

        headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
        header_to_col_index = {
            str(header).strip(): idx
            for idx, header in enumerate(headers, 1)
            if header is not None and str(header).strip()
        }

        if output_dir is None:
            output_dir = tempfile.gettempdir()
        output_dir = os.path.normpath(output_dir)
        os.makedirs(output_dir, exist_ok=True)


        def _write_cell_safe(row_num: int, col_num: int, value: Any) -> None:
            cell = ws.cell(row=row_num, column=col_num)
            if cell.__class__.__name__ != "MergedCell":
                cell.value = value
                return

            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row <= row_num <= merged_range.max_row and merged_range.min_col <= col_num <= merged_range.max_col:
                    ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                    return

        # start write row: first row under header that has any mapped-column value, else header+1
        mapped_col_indices = [
            header_to_col_index.get(str(excel_col).strip())
            for excel_col, data_col in column_mappings.items()
            if data_col
        ]
        mapped_col_indices = [c for c in mapped_col_indices if c]

        data_row = header_row + 1
        for r in range(header_row + 1, min(ws.max_row + 1, header_row + 5000)):
            has_data = False
            for c in mapped_col_indices or [1]:
                if ws.cell(row=r, column=c).value not in (None, ""):
                    has_data = True
                    break
            if has_data:
                data_row = r
                break

        logger.info(f"Detected header row={header_row}, write_start_row={data_row}")
        logger.info(f"Header columns at row {header_row}: {[str(h).strip() if h is not None else None for h in headers]}")
        logger.info(f"Mapped column indexes: {mapped_col_indices}")

        # clear only mapped columns in old data region, keep layout/other formula/style
        if mapped_col_indices:
            max_clear_row = min(ws.max_row, data_row + max(len(data), 2000))
            for r in range(data_row, max_clear_row + 1):
                for c in mapped_col_indices:
                    _write_cell_safe(r, c, None)

        # Fill data
        for idx, row_data in enumerate(data):
            row_num = data_row + idx

            # Map theo format: {"<ten_cot_excel>": "<ten_cot_sql_or_db> | null"}
            for excel_col, data_col in column_mappings.items():
                if not data_col:
                    continue
                col_index = header_to_col_index.get(str(excel_col).strip())
                if col_index and data_col in row_data:
                    _write_cell_safe(row_num, col_index, row_data[data_col])

            # Auto-map fallback theo trùng tên header với key row_data
            for col_idx, header in enumerate(headers, 1):
                if header and header in row_data and ws.cell(row=row_num, column=col_idx).value is None:
                    _write_cell_safe(row_num, col_idx, row_data[header])

            if idx < 3:
                row_preview = [ws.cell(row=row_num, column=c).value for c in range(1, min(ws.max_column, 18) + 1)]
                logger.info(f"📝 Written row {row_num} preview: {row_preview}")
        
        # Save output
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"filled_{timestamp}.xlsx"
        output_path = os.path.join(output_dir, filename)
        
        wb.save(output_path)
        wb.close()
        
        logger.info(f"✅ Saved filled file: {output_path}")
        return output_path

    def _save_as_json(self, template_path: str, data: list[dict], output_dir: str = None) -> str:
        """Fallback: Save as JSON if openpyxl not available."""
        if output_dir is None:
            output_dir = "/tmp"
        
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"data_{timestamp}.json"
        output_path = os.path.join(output_dir, filename)
        
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return output_path


# Singleton
_file_data_filler: FileDataFillerService | None = None


def get_file_data_filler() -> FileDataFillerService:
    global _file_data_filler
    if _file_data_filler is None:
        _file_data_filler = FileDataFillerService()
    return _file_data_filler