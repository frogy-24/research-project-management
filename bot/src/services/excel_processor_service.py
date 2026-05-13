"""
ExcelProcessorService - Đọc và phân tích cấu trúc file Excel
"""

import os
import json
from typing import Any
from pathlib import Path

from src.utilities import get_logger

logger = get_logger(__name__)


class ExcelProcessorService:
    """Service để đọc và phân tích file Excel."""

    @staticmethod
    def is_valid_excel(file_path: str) -> bool:
        """Kiểm tra file có phải Excel không."""
        valid_extensions = {".xlsx", ".xls"}
        ext = Path(file_path).suffix.lower()
        return ext in valid_extensions

    @staticmethod
    def get_file_extension(file_path: str) -> str:
        """Lấy đuôi file."""
        return Path(file_path).suffix.lower()

    @classmethod
    def analyze_excel(cls, file_path: str) -> dict[str, Any]:
        """
        Phân tích file Excel và trả về cấu trúc.
        
        Returns:
            {
                "fileName": str,
                "fileExtension": str,
                "sheetNames": [str],
                "structure": {
                    "sheet_name": {
                        "rowCount": int,
                        "columnCount": int,
                        "headers": [str],
                        "sampleData": [[any]],
                        "styles": {
                            "headerFont": str,
                            "headerFill": str,
                            "columnWidths": [float]
                        }
                    }
                },
                "columnMappings": {
                    "col_index": {
                        "name": str,
                        "sampleValues": [any],
                        "dataType": str  # "text", "number", "date", "mixed"
                    }
                }
            }
        """
        if not cls.is_valid_excel(file_path):
            raise ValueError(f"File không phải Excel: {file_path}")

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed")
            return cls._get_basic_info(file_path)

        wb = load_workbook(file_path, data_only=True)
        
        result = {
            "fileName": os.path.basename(file_path),
            "fileExtension": cls.get_file_extension(file_path),
            "sheetNames": wb.sheetnames,
            "structure": {},
            "columnMappings": {}
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = []
            sample_data = []
            
            # Đọc headers (dòng đầu tiên)
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")
            
            # Đọc sample data (5 dòng đầu tiên sau header)
            for row in range(2, min(7, ws.max_row + 1)):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    row_data.append(ws.cell(row=row, column=col).value)
                sample_data.append(row_data)
            
            # Thu thập styles
            styles = cls._extract_styles(ws, headers)
            
            # Column mappings - phân tích từng cột
            column_mappings = {}
            for col_idx, header in enumerate(headers, 1):
                col_values = []
                for row in range(2, min(20, ws.max_row + 1)):  # Sample 20 rows
                    val = ws.cell(row=row, column=col_idx).value
                    if val is not None:
                        col_values.append(val)
                
                column_mappings[col_idx] = {
                    "name": header,
                    "sampleValues": col_values[:5],
                    "dataType": cls._infer_data_type(col_values)
                }
            
            result["structure"][sheet_name] = {
                "rowCount": ws.max_row,
                "columnCount": ws.max_column,
                "headers": headers,
                "sampleData": sample_data,
                "styles": styles
            }
            result["columnMappings"].update(column_mappings)

        wb.close()
        logger.info(f"✅ Analyzed Excel: {result['fileName']}, sheets: {result['sheetNames']}")
        return result

    @staticmethod
    def _extract_styles(ws, headers: list[str]) -> dict[str, Any]:
        """Trích xuất styles từ worksheet."""
        styles = {
            "headerFont": [],
            "headerFills": [],
            "columnWidths": []
        }
        
        try:
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return styles
        
        # Header styles
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            if cell.font and cell.font.bold:
                styles["headerFont"].append("bold")
            else:
                styles["headerFont"].append("normal")
            
            if cell.fill and hasattr(cell.fill, 'fgColor'):
                styles["headerFills"].append(str(cell.fill.fgColor.rgb) if cell.fill.fgColor else None)
        
        # Column widths
        for col in range(1, len(headers) + 1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            if col_letter in ws.column_dimensions:
                styles["columnWidths"].append(ws.column_dimensions[col_letter].width)
            else:
                styles["columnWidths"].append(10)  # default
        
        return styles

    @staticmethod
    def _infer_data_type(values: list[Any]) -> str:
        """Infer data type from sample values."""
        if not values:
            return "unknown"
        
        # Check if all numeric
        numeric_count = 0
        date_count = 0
        text_count = 0
        
        for val in values:
            val_str = str(val).strip()
            
            # Check numeric
            try:
                float(val_str.replace(",", ""))
                numeric_count += 1
                continue
            except (ValueError, AttributeError):
                pass
            
            # Check date patterns
            import re
            if re.match(r'\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', val_str):
                date_count += 1
                continue
            
            # Check if mostly text
            if len(val_str) > 3:
                text_count += 1
        
        total = len(values)
        if numeric_count / total > 0.7:
            return "number"
        elif date_count / total > 0.5:
            return "date"
        elif text_count / total > 0.5:
            return "text"
        return "mixed"

    @staticmethod
    def _get_basic_info(file_path: str) -> dict[str, Any]:
        """Fallback khi không có openpyxl."""
        return {
            "fileName": os.path.basename(file_path),
            "fileExtension": Path(file_path).suffix,
            "sheetNames": [],
            "structure": {},
            "columnMappings": {},
            "error": "openpyxl not available"
        }

    @classmethod
    def describe_for_llm(cls, analysis: dict[str, Any]) -> str:
        """
        Tạo mô tả text từ analysis để đưa vào LLM prompt.
        
        Format:
        ```
        File: filename.xlsx
        Sheets: [sheet1, sheet2]
        
        Sheet: sheet1
        - Total rows: N, columns: M
        - Headers: [col1, col2, col3, ...]
        - Sample data:
          [row1 data]
          [row2 data]
        - Column details:
          - Column 1 (col1): text - sample: ["value1", "value2"]
          - Column 2 (col2): number - sample: [1, 2, 3]
        ```
        """
        lines = []
        lines.append(f"File: {analysis['fileName']}")
        lines.append(f"Sheets: {analysis['sheetNames']}")
        
        for sheet_name, structure in analysis.get("structure", {}).items():
            lines.append(f"\nSheet: {sheet_name}")
            lines.append(f"- Total rows: {structure.get('rowCount', 0)}, columns: {structure.get('columnCount', 0)}")
            lines.append(f"- Headers: {structure.get('headers', [])}")
            
            sample = structure.get("sampleData", [])
            if sample:
                lines.append(f"- Sample data (first {len(sample)} rows):")
                for row in sample:
                    lines.append(f"  {row}")
            
            # Column mappings
            column_mappings = analysis.get("columnMappings", {})
            lines.append("- Column details:")
            for col_idx, mapping in sorted(column_mappings.items()):
                name = mapping.get("name", f"Column_{col_idx}")
                dtype = mapping.get("dataType", "unknown")
                samples = mapping.get("sampleValues", [])
                lines.append(f"  - {name}: {dtype} - sample: {samples[:3]}")
        
        return "\n".join(lines)


# Singleton
_excel_processor: ExcelProcessorService | None = None


def get_excel_processor() -> ExcelProcessorService:
    global _excel_processor
    if _excel_processor is None:
        _excel_processor = ExcelProcessorService()
    return _excel_processor