"""
Report Generation Consumer - Tao bao cao voi RabbitMQ + LLM
"""

import asyncio
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import aio_pika

if str(Path(__file__).resolve().parents[2]) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.db import db
from src.services.llm_service import LLMService
from src.utilities import get_logger

logger = get_logger(__name__)

RABBITMQ_URL = os.getenv("RABBITMQ_URL", "amqp://admin:admin123@localhost:5672/")
QUEUE_NAME = os.getenv("REPORT_QUEUE", "report_generation_queue")
PREFETCH_COUNT = int(os.getenv("REPORT_PREFETCH", "1"))
REPORT_OUTPUT_DIR = os.getenv("REPORT_OUTPUT_DIR", "/home/caoviet/Documents/reports_shared")
# Next.js server base URL for fetching uploaded files
NEXTJS_BASE_URL = os.getenv("NEXTJS_BASE_URL", "http://localhost:3000")


async def _update_job_status(
    job_id: str,
    status: str,
    progress: int = 0,
    result_url: str | None = None,
    error: str | None = None,
) -> None:
    now = datetime.now()
    set_parts = ['"status" = $1', '"progress" = $2', '"updatedAt" = $3']
    params: list[Any] = [status, progress, now]
    param_idx = 4

    if result_url:
        set_parts.append(f'"resultUrl" = ${param_idx}')
        params.append(result_url)
        param_idx += 1

    if error:
        set_parts.append(f'"error" = ${param_idx}')
        params.append(error)
        param_idx += 1

    if status in {"COMPLETED", "FAILED"}:
        set_parts.append(f'"completedAt" = ${param_idx}')
        params.append(now)
        param_idx += 1

    params.append(job_id)
    query = f"""
        UPDATE "ReportJob"
        SET {', '.join(set_parts)}
        WHERE id = ${param_idx}
    """
    await db.execute(query, *params)


async def _generate_report_with_llm(
    report_type: str,
    data: dict[str, Any],
    llm_service: LLMService,
) -> str:
    """Generate report content using LLM"""
    
    prompt = f"""
Ban la mot chuyen gia tao bao cao. Tao mot bao cao {report_type} chi tiet.

Du lieu:
{json.dumps(data, ensure_ascii=False, indent=2)}

Yeu cau:
1. Dinh dang bang bang cach su dung Markdown table
2. Tong hop so lieu chinh xac
3. Tinh toan cac chi so thong ke neu co
4. Dua ra nhan xet neu thich hop

Chi tra ve noi dung Markdown cua bao cao, khong giai thich them.
"""
    
    messages = [
        {"role": "system", "content": "Ban la mot chuyen gia tao bao cao chuyen nghiep. Tra loi chi bang noi dung bao cao Markdown, khong giai thich them."},
        {"role": "user", "content": prompt}
    ]
    
    return await llm_service.chat_completion(messages)


async def _process_payload(payload: dict[str, Any]) -> dict[str, Any]:
    report_type = payload.get("reportType")
    template_url = payload.get("templateUrl")
    template_type = payload.get("templateType")
    job_id = payload.get("jobId")

    if not report_type:
        return {"error": "Missing reportType"}

    if job_id:
        await _update_job_status(job_id, "PROCESSING", progress=10)

    # Check if using uploaded template
    if template_type == "uploaded" and template_url:
        # For uploaded templates - download from Next.js, fill with data, upload
        if job_id:
            await _update_job_status(job_id, "PROCESSING", progress=30)
        
        # Download template from Next.js
        temp_path, template_filename = await _download_template_from_nextjs(template_url)
        logger.info(f"Downloaded template: {temp_path}")
        
        # Fetch data for filling template
        parameters = payload.get("parameters") or {}
        parameters["callRoundId"] = payload.get("callRoundId")
        report_data = await _fetch_report_data(report_type, parameters)
        
        if job_id:
            await _update_job_status(job_id, "PROCESSING", progress=60)
        
        # Fill Excel template with data
        output_path = await _fill_excel_template(temp_path, report_data, template_filename)
        logger.info(f"Filled template saved to: {output_path}")
        
        # Upload to Next.js reports directory
        result_url = await _upload_report_to_nextjs(output_path, report_type)
        
        if job_id:
            await _update_job_status(job_id, "COMPLETED", progress=100, result_url=result_url)
        
        return {"resultUrl": result_url}

    # For auto-generated reports - use LLM
    parameters = payload.get("parameters") or {}
    data = await _fetch_report_data(report_type, parameters)
    
    if job_id:
        await _update_job_status(job_id, "PROCESSING", progress=50)

    # Generate report with LLM
    llm_service = LLMService()
    report_content = await _generate_report_with_llm(report_type, data, llm_service)

    if job_id:
        await _update_job_status(job_id, "PROCESSING", progress=80)

    # Save report to temp file then upload to Next.js
    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    filename = f"report_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md"
    filepath = os.path.join(REPORT_OUTPUT_DIR, filename)
    
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f"# Bao cao: {report_type}\n")
        f.write(f"Ngay tao: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(report_content)

    # Upload to Next.js server
    result_url = await _upload_report_to_nextjs(filepath, report_type)

    if job_id:
        await _update_job_status(job_id, "COMPLETED", progress=100, result_url=result_url)

    return {"resultUrl": result_url, "content": report_content}


async def _upload_report_to_nextjs(filepath: str, report_type: str | None = None) -> str:
    """Upload generated report file to Next.js server"""
    import aiohttp
    
    filename = os.path.basename(filepath)
    upload_url = f"{NEXTJS_BASE_URL}/api/reports/upload"
    logger.info(f"Uploading report to: {upload_url}")
    
    try:
        # Read file content first
        with open(filepath, 'rb') as f:
            file_content = f.read()
        
        form_data = aiohttp.FormData()
        form_data.add_field('file', file_content, filename=filename, content_type='application/octet-stream')
        if report_type:
            form_data.add_field('reportType', report_type)
        
        headers = {"X-Internal-Request": "true"}
        async with aiohttp.ClientSession() as session:
            async with session.post(upload_url, data=form_data, headers=headers) as response:
                if response.status != 200:
                    text = await response.text()
                    raise Exception(f"Failed to upload report: {response.status} {text}")
                
                result = await response.json()
                logger.info(f"Uploaded report, URL: {result.get('url')}")
                return result.get('url', f"/uploads/reports/{filename}")
                
    except aiohttp.ClientError as e:
        logger.error(f"HTTP error uploading report: {e}")
        raise Exception(f"Failed to upload report to Next.js: {e}")


async def _download_template_from_nextjs(template_url: str) -> tuple[str, str]:
    """Download template from Next.js server and save to temp file"""
    import aiohttp
    
    # template_url comes as '/uploads/templates/filename.xlsx'
    template_filename = os.path.basename(template_url)
    
    # Build full URL for Next.js server
    full_url = f"{NEXTJS_BASE_URL}{template_url}"
    logger.info(f"Downloading template from: {full_url}")
    
    # Download to temp file
    temp_path = os.path.join("/tmp", f"template_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{template_filename}")
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(full_url) as response:
                if response.status != 200:
                    raise FileNotFoundError(f"Failed to download template: {response.status} {full_url}")
                
                with open(temp_path, 'wb') as f:
                    async for chunk in response.content.iter_chunked(8192):
                        f.write(chunk)
        
        logger.info(f"Downloaded template to: {temp_path}")
        return temp_path, template_filename
        
    except aiohttp.ClientError as e:
        logger.error(f"HTTP error downloading template: {e}")
        raise FileNotFoundError(f"Failed to download template from {full_url}: {e}")


async def _fetch_report_data(report_type: str, parameters: dict[str, Any]) -> dict[str, Any]:
    """Fetch data for report based on type - includes projects, registrations, and statistics"""
    
    call_round_id = parameters.get("callRoundId")
    from_date = parameters.get("fromDate")
    to_date = parameters.get("toDate")
    department_id = parameters.get("departmentId")

    result_data = {
        "reportType": report_type,
        "parameters": parameters,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
    }

    # ========== 1. FETCH CALL ROUNDS ==========
    call_round_filter = ""
    cr_params = []
    cr_idx = 1
    if call_round_id:
        call_round_filter = f"WHERE id = ${cr_idx}"
        cr_params = [call_round_id]
    
    try:
        call_rounds = await db.fetch_all(f'''
            SELECT id, name, "registrationStartDate", "registrationEndDate", 
                   "projectStartDate", "projectEndDate", defense_date, is_active
            FROM "CallRound"
            {call_round_filter}
            ORDER BY "registrationStartDate" DESC
            LIMIT 50
        ''', *cr_params)
        result_data["callRounds"] = [
            {
                "id": r["id"],
                "name": r["name"],
                "registrationStartDate": str(r["registrationStartDate"]) if r["registrationStartDate"] else None,
                "registrationEndDate": str(r["registrationEndDate"]) if r["registrationEndDate"] else None,
                "projectStartDate": str(r["projectStartDate"]) if r["projectStartDate"] else None,
                "projectEndDate": str(r["projectEndDate"]) if r["projectEndDate"] else None,
                "isActive": r.get("is_active"),
            }
            for r in call_rounds
        ]
    except Exception as e:
        logger.error(f"Error fetching call rounds: {e}")
        result_data["callRounds"] = []

    # ========== 2. FETCH PROJECT REGISTRATIONS ==========
    reg_params = []
    reg_idx = 1
    reg_filters = []
    
    if call_round_id:
        reg_filters.append(f'pr."callRoundId" = ${reg_idx}')
        reg_params.append(call_round_id)
        reg_idx += 1

    if from_date:
        reg_filters.append(f'cr."registrationStartDate" >= ${reg_idx}')
        reg_params.append(from_date)
        reg_idx += 1

    if to_date:
        reg_filters.append(f'cr."registrationStartDate" <= ${reg_idx}')
        reg_params.append(to_date)
        reg_idx += 1

    reg_where = "WHERE " + " AND ".join(reg_filters) if reg_filters else ""
    
    try:
        registrations = await db.fetch_all(f'''
            SELECT 
                pr.id,
                pr.title,
                pr.status,
                pr."createdAt",
                u.id as student_id,
                u.name as student_name,
                u.email as student_email,
                u.code as student_code,
                d.id as department_id,
                d.name as department_name,
                m.name as major_name,
                cr.name as call_round_name,
                ins.name as instructor_name
            FROM "ProjectRegistration" pr
            JOIN "User" u ON pr."userId" = u.id
            LEFT JOIN "CallRound" cr ON pr."callRoundId" = cr.id
            LEFT JOIN "Department" d ON u."departmentId" = d.id
            LEFT JOIN "Major" m ON u."majorId" = m.id
            LEFT JOIN "User" ins ON pr."instructorId" = ins.id
            {reg_where}
            ORDER BY pr."createdAt" DESC
            LIMIT 500
        ''', *reg_params)
        result_data["registrations"] = [
            {
                "id": r["id"],
                "title": r["title"],
                "status": r["status"],
                "studentId": r["student_id"],
                "studentName": r["student_name"],
                "studentEmail": r["student_email"],
                "studentCode": r["student_code"],
                "departmentId": r["department_id"],
                "department": r["department_name"],
                "major": r["major_name"],
                "callRound": r["call_round_name"],
                "instructorName": r["instructor_name"],
                "createdAt": str(r["createdAt"]) if r["createdAt"] else None,
            }
            for r in registrations
        ]
    except Exception as e:
        logger.error(f"Error fetching registrations: {e}")
        result_data["registrations"] = []

    # ========== 3. FETCH PROJECTS ==========
    proj_params = []
    proj_idx = 1
    proj_filters = []
    
    if call_round_id:
        proj_filters.append(f'"callRoundId" = ${proj_idx}')
        proj_params.append(call_round_id)
        proj_idx += 1
    
    proj_where = "WHERE " + " AND ".join(proj_filters) if proj_filters else ""
    
    try:
        projects = await db.fetch_all(f'''
            SELECT 
                p.id,
                p.code,
                p.title,
                p.status,
                p."budgetRequested",
                p."budgetApproved",
                p."createdAt",
                u.name as leader_name,
                d.name as department_name,
                cr.name as call_round_name
            FROM "Project" p
            LEFT JOIN "User" u ON p."leaderId" = u.id
            LEFT JOIN "Department" d ON u."departmentId" = d.id
            LEFT JOIN "CallRound" cr ON p."callRoundId" = cr.id
            {proj_where}
            ORDER BY p."createdAt" DESC
            LIMIT 500
        ''', *proj_params)
        result_data["projects"] = [
            {
                "id": proj["id"],
                "code": proj["code"],
                "title": proj["title"],
                "status": proj["status"],
                "budgetRequested": float(proj["budgetRequested"]) if proj["budgetRequested"] else 0,
                "budgetApproved": float(proj["budgetApproved"]) if proj["budgetApproved"] else 0,
                "leaderName": proj["leader_name"],
                "department": proj["department_name"],
                "callRound": proj["call_round_name"],
                "createdAt": str(proj["createdAt"]) if proj["createdAt"] else None,
            }
            for proj in projects
        ]
    except Exception as e:
        logger.error(f"Error fetching projects: {e}")
        result_data["projects"] = []

    # ========== 4. FETCH USERS ==========
    try:
        users = await db.fetch_all('''
            SELECT id, code, name, email, role, "departmentId"
            FROM "User"
            ORDER BY "createdAt" DESC
            LIMIT 200
        ''')
        result_data["users"] = [
            {
                "id": u["id"],
                "code": u["code"],
                "name": u["name"],
                "email": u["email"],
                "role": u["role"],
                "departmentId": u["departmentId"],
            }
            for u in users
        ]
    except Exception as e:
        logger.error(f"Error fetching users: {e}")
        result_data["users"] = []

    # ========== 5. COMPUTE STATISTICS ==========
    total_registrations = len(result_data["registrations"])
    total_projects = len(result_data["projects"])
    total_users = len(result_data["users"])

    # Registration stats by status
    reg_by_status = {}
    for reg in result_data["registrations"]:
        status = reg.get("status", "UNKNOWN")
        reg_by_status[status] = reg_by_status.get(status, 0) + 1

    # Project stats by status
    proj_by_status = {}
    total_budget_requested = 0
    total_budget_approved = 0
    for proj in result_data["projects"]:
        status = proj.get("status", "UNKNOWN")
        proj_by_status[status] = proj_by_status.get(status, 0) + 1
        total_budget_requested += proj.get("budgetRequested", 0)
        total_budget_approved += proj.get("budgetApproved", 0)

    # User stats by role
    user_by_role = {}
    for user in result_data["users"]:
        role = user.get("role", "UNKNOWN")
        user_by_role[role] = user_by_role.get(role, 0) + 1

    # Stats by department
    reg_by_dept = {}
    proj_by_dept = {}
    for reg in result_data["registrations"]:
        dept = reg.get("department") or "Unknown"
        reg_by_dept[dept] = reg_by_dept.get(dept, 0) + 1
    for proj in result_data["projects"]:
        dept = proj.get("department") or "Unknown"
        proj_by_dept[dept] = proj_by_dept.get(dept, 0) + 1

    result_data["statistics"] = {
        "totalRegistrations": total_registrations,
        "totalProjects": total_projects,
        "totalUsers": total_users,
        "registrationByStatus": reg_by_status,
        "projectByStatus": proj_by_status,
        "userByRole": user_by_role,
        "registrationByDepartment": reg_by_dept,
        "projectByDepartment": proj_by_dept,
        "totalBudgetRequested": total_budget_requested,
        "totalBudgetApproved": total_budget_approved,
        "callRoundsCount": len(result_data["callRounds"]),
    }

    logger.info(f"Report data fetched: {total_registrations} registrations, {total_projects} projects, {total_users} users")
    return result_data


def _fill_excel_template(template_path: str, data: dict[str, Any], original_filename: str) -> str:
    """Fill Excel template with report data"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        import copy
    except ImportError:
        logger.warning("openpyxl not installed, saving raw data as JSON instead")
        # Fallback: save as JSON
        output_path = template_path.replace(".xlsx", "_data.json").replace("/tmp/", REPORT_OUTPUT_DIR + "/")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return output_path
    
    wb = load_workbook(template_path)
    
    stats = data.get("statistics", {})
    registrations = data.get("registrations", [])
    projects = data.get("projects", [])
    call_rounds = data.get("callRounds", [])
    
    # Find sheet named "Thong_ke" or first sheet
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if "thong_ke" in sheet_name.lower() or "stats" in sheet_name.lower():
            target_sheet = wb[sheet_name]
            break
    
    if target_sheet is None:
        target_sheet = wb.active
    
    logger.info(f"Filling sheet: {target_sheet.title}")
    
    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row_idx = 1
    
    # Title
    target_sheet.cell(row=row_idx, column=1, value=f"BÁO CÁO THỐNG KÊ - {data.get('reportType', 'Tổng hợp')}")
    target_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    row_idx += 1
    
    # Generated date
    target_sheet.cell(row=row_idx, column=1, value=f"Ngày tạo: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    row_idx += 2
    
    # ========== SUMMARY SECTION ==========
    target_sheet.cell(row=row_idx, column=1, value="TỔNG QUAN")
    target_sheet.cell(row=row_idx, column=1).font = header_font
    row_idx += 1
    
    # Total stats
    summary_data = [
        ("Tổng số đăng ký", stats.get("totalRegistrations", 0)),
        ("Tổng số đề tài", stats.get("totalProjects", 0)),
        ("Tổng số người dùng", stats.get("totalUsers", 0)),
        ("Tổng số đợt đăng ký", stats.get("callRoundsCount", 0)),
        ("Ngân sách yêu cầu (VNĐ)", stats.get("totalBudgetRequested", 0)),
        ("Ngân sách duyệt (VNĐ)", stats.get("totalBudgetApproved", 0)),
    ]
    
    for label, value in summary_data:
        target_sheet.cell(row=row_idx, column=1, value=label)
        target_sheet.cell(row=row_idx, column=2, value=value)
        row_idx += 1
    
    row_idx += 1
    
    # ========== REGISTRATION BY STATUS ==========
    if stats.get("registrationByStatus"):
        target_sheet.cell(row=row_idx, column=1, value="ĐĂNG KÝ THEO TRẠNG THÁI")
        target_sheet.cell(row=row_idx, column=1).font = header_font
        row_idx += 1
        
        target_sheet.cell(row=row_idx, column=1, value="Trạng thái")
        target_sheet.cell(row=row_idx, column=2, value="Số lượng")
        for col in [1, 2]:
            target_sheet.cell(row=row_idx, column=col).font = header_font
            target_sheet.cell(row=row_idx, column=col).fill = header_fill
            target_sheet.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
        
        for status, count in stats["registrationByStatus"].items():
            target_sheet.cell(row=row_idx, column=1, value=status)
            target_sheet.cell(row=row_idx, column=2, value=count)
            for col in [1, 2]:
                target_sheet.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1
        
        row_idx += 1
    
    # ========== PROJECT BY STATUS ==========
    if stats.get("projectByStatus"):
        target_sheet.cell(row=row_idx, column=1, value="ĐỀ TÀI THEO TRẠNG THÁI")
        target_sheet.cell(row=row_idx, column=1).font = header_font
        row_idx += 1
        
        target_sheet.cell(row=row_idx, column=1, value="Trạng thái")
        target_sheet.cell(row=row_idx, column=2, value="Số lượng")
        for col in [1, 2]:
            target_sheet.cell(row=row_idx, column=col).font = header_font
            target_sheet.cell(row=row_idx, column=col).fill = header_fill
            target_sheet.cell(row=row_idx, column=col).border = thin_border
        row_idx += 1
        
        for status, count in stats["projectByStatus"].items():
            target_sheet.cell(row=row_idx, column=1, value=status)
            target_sheet.cell(row=row_idx, column=2, value=count)
            for col in [1, 2]:
                target_sheet.cell(row=row_idx, column=col).border = thin_border
            row_idx += 1
        
        row_idx += 1
    
    # ========== REGISTRATION DETAILS ==========
    if registrations:
        target_sheet.cell(row=row_idx, column=1, value="CHI TIẾT ĐĂNG KÝ")
        target_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        row_idx += 1
        
        # Headers
        headers = ["STT", "Tên đề tài", "Sinh viên", "Email", "Khoa", "Ngành", "Giảng viên", "Trạng thái", "Ngày tạo"]
        for col, header in enumerate(headers, 1):
            cell = target_sheet.cell(row=row_idx, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row_idx += 1
        
        # Data rows (limit to first 100)
        for idx, reg in enumerate(registrations[:100], 1):
            row_data = [
                idx,
                reg.get("title", ""),
                reg.get("studentName", ""),
                reg.get("studentEmail", ""),
                reg.get("department", ""),
                reg.get("major", ""),
                reg.get("instructorName", ""),
                reg.get("status", ""),
                reg.get("createdAt", "")[:10] if reg.get("createdAt") else "",
            ]
            for col, value in enumerate(row_data, 1):
                cell = target_sheet.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
            row_idx += 1
        
        row_idx += 1
    
    # ========== PROJECT DETAILS ==========
    if projects:
        target_sheet.cell(row=row_idx, column=1, value="CHI TIẾT ĐỀ TÀI")
        target_sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
        row_idx += 1
        
        # Headers
        proj_headers = ["STT", "Mã", "Tên đề tài", "Trưởng nhóm", "Khoa", "Ngân sách YC", "Ngân sách duyệt", "Trạng thái"]
        for col, header in enumerate(proj_headers, 1):
            cell = target_sheet.cell(row=row_idx, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row_idx += 1
        
        # Data rows (limit to first 100)
        for idx, proj in enumerate(projects[:100], 1):
            row_data = [
                idx,
                proj.get("code", ""),
                proj.get("title", ""),
                proj.get("leaderName", ""),
                proj.get("department", ""),
                proj.get("budgetRequested", 0),
                proj.get("budgetApproved", 0),
                proj.get("status", ""),
            ]
            for col, value in enumerate(row_data, 1):
                cell = target_sheet.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
            row_idx += 1
    
    # Adjust column widths
    for column in target_sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        target_sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to output
    output_filename = original_filename.replace(".xlsx", f"_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    output_path = os.path.join(REPORT_OUTPUT_DIR, output_filename)
    os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
    wb.save(output_path)
    
    logger.info(f"Saved filled Excel report to: {output_path}")
    return output_path


async def _send_reply(
    channel: aio_pika.Channel,
    reply_to: str,
    correlation_id: str | None,
    payload: dict[str, Any],
) -> None:
    message = aio_pika.Message(
        body=json.dumps(payload).encode("utf-8"),
        content_type="application/json",
        correlation_id=correlation_id,
        delivery_mode=aio_pika.DeliveryMode.PERSISTENT,
    )
    await channel.default_exchange.publish(message, routing_key=reply_to)


async def _handle_message(message: aio_pika.IncomingMessage) -> None:
    async with message.process(ignore_processed=True):
        try:
            payload = json.loads(message.body.decode("utf-8"))
        except json.JSONDecodeError:
            logger.error("Invalid JSON payload")
            return

        logger.info(f"Processing report payload: {payload}")

        try:
            result = await _process_payload(payload)
        except Exception as exc:
            job_id = payload.get("jobId")
            if job_id:
                await _update_job_status(job_id, "FAILED", error=str(exc))
            raise

        result_payload = {
            "success": "error" not in result,
            "data": result if "error" not in result else None,
            "error": result.get("error"),
        }

        if message.reply_to:
            await _send_reply(
                channel=message.channel,
                reply_to=message.reply_to,
                correlation_id=message.correlation_id,
                payload=result_payload,
            )


async def main() -> None:
    logger.info("Starting Report Generation Consumer...")
    logger.info(f"RabbitMQ URL: {RABBITMQ_URL}")
    logger.info(f"Queue: {QUEUE_NAME}")

    connection = await aio_pika.connect_robust(RABBITMQ_URL)
    async with connection:
        await db.connect()
        channel = await connection.channel()
        await channel.set_qos(prefetch_count=PREFETCH_COUNT)

        queue = await channel.declare_queue(QUEUE_NAME, durable=True)

        logger.info("Consumer ready. Waiting for messages...")
        await queue.consume(_handle_message)

        try:
            await asyncio.Future()
        finally:
            await db.close()


if __name__ == "__main__":
    asyncio.run(main())
