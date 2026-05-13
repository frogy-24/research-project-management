"""
File Processor API Route - Upload file Excel, phân tích + fill data
"""

import os
import sys
from pathlib import Path
from typing import Any

from fastapi import APIRouter, UploadFile, File, Form, HTTPException

# Add src to path
if str(Path(__file__).resolve().parents[2]) not in sys.path:
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from src.services.file_data_filler_service import get_file_data_filler
from src.utilities import get_logger, log_async_execution

logger = get_logger(__name__)
router = APIRouter(prefix="/file-processor", tags=["File Processor"])

TEMP_UPLOAD_DIR = os.getenv("TEMP_UPLOAD_DIR", "/tmp/uploads")
REPORT_OUTPUT_DIR = os.getenv("REPORT_OUTPUT_DIR", "/home/caoviet/Documents/reports_shared")


@router.post("/analyze")
@log_async_execution
async def analyze_file(
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """
    Phân tích file Excel và trả về cấu trúc.
    
    Upload file Excel -> Trả về thông tin cấu trúc file
    """
    logger.info(f"📤 Analyzing uploaded file: {file.filename}")
    
    # Validate file type
    allowed_extensions = {".xlsx", ".xls"}
    file_ext = Path(file.filename).suffix.lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Chỉ hỗ trợ file Excel: {allowed_extensions}"
        )
    
    # Save uploaded file to temp
    os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_UPLOAD_DIR, f"upload_{file.filename}")
    
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
        
        # Analyze file
        from src.services.excel_processor_service import get_excel_processor
        processor = get_excel_processor()
        analysis = processor.analyze_excel(temp_path)
        description = processor.describe_for_llm(analysis)
        
        return {
            "success": True,
            "fileName": file.filename,
            "fileSize": len(content),
            "analysis": analysis,
            "description": description
        }
        
    except Exception as e:
        logger.error(f"❌ Analyze file error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        # Cleanup temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)


@router.post("/fill")
@log_async_execution
async def fill_file_data(
    file: UploadFile = File(...),
    output_format: str = Form("xlsx"),  # xlsx, json
) -> dict[str, Any]:
    """
    Upload file Excel -> LLM phân tích -> Lấy data từ DB -> Fill vào file -> Trả về file đã fill.
    
    Flow:
    1. User upload file Excel (có headers nhưng chưa có data)
    2. LLM phân tích file: xác định đây là danh sách đăng ký đề tài, hội đồng, etc.
    3. LLM sinh SQL để lấy data phù hợp
    4. Fill data vào file
    5. Trả về file đã fill
    """
    logger.info(f"📤 Processing file for data fill: {file.filename}")
    
    # Validate file type
    allowed_extensions = {".xlsx", ".xls"}
    file_ext = Path(file.filename).suffix.lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Chỉ hỗ trợ file Excel: {allowed_extensions}"
        )
    
    # Save uploaded file to temp
    os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_UPLOAD_DIR, f"upload_{file.filename}")
    
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
        
        # Process file
        filler = get_file_data_filler()
        result = await filler.analyze_and_fill(temp_path, REPORT_OUTPUT_DIR)
        
        if result["success"]:
            # Upload filled file to Next.js
            result_url = await _upload_to_nextjs(result["outputPath"], file.filename)
            
            return {
                "success": True,
                "fileName": file.filename,
                "resultUrl": result_url,
                "analysis": result.get("analysis"),
                "rowCount": result.get("rowCount", 0),
                "message": result.get("message")
            }
        else:
            raise HTTPException(status_code=500, detail=result.get("error"))
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Fill file error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        # Cleanup temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)


async def _upload_to_nextjs(filepath: str, original_filename: str) -> str:
    """Upload file đã fill đến Next.js server."""
    import aiohttp
    
    nextjs_url = os.getenv("NEXTJS_BASE_URL", "http://localhost:3000")
    upload_url = f"{nextjs_url}/api/reports/upload"
    
    filename = os.path.basename(filepath)
    
    try:
        with open(filepath, 'rb') as f:
            file_content = f.read()
        
        form_data = aiohttp.FormData()
        form_data.add_field('file', file_content, filename=filename, content_type='application/octet-stream')
        
        headers = {"X-Internal-Request": "true"}
        
        async with aiohttp.ClientSession() as session:
            async with session.post(upload_url, data=form_data, headers=headers) as response:
                if response.status != 200:
                    text = await response.text()
                    raise Exception(f"Upload failed: {response.status} {text}")
                
                result = await response.json()
                return result.get('url', f"/uploads/reports/{filename}")
                
    except aiohttp.ClientError as e:
        logger.error(f"Upload error: {e}")
        return f"/uploads/reports/{filename}"  # Fallback URL


@router.post("/quick-fill")
@log_async_execution  
async def quick_fill(
    file: UploadFile = File(...),
    sql: str = Form(...),  # SQL query do user cung cap
) -> dict[str, Any]:
    """
    Upload file Excel + SQL query -> Fill data -> Trả về file.
    
    Dùng khi user tự viết SQL hoặc muốn dùng lại SQL cũ.
    """
    logger.info(f"📤 Quick fill with custom SQL: {file.filename}")
    
    # Validate
    allowed_extensions = {".xlsx", ".xls"}
    file_ext = Path(file.filename).suffix.lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(status_code=400, detail="Chỉ hỗ trợ file Excel")
    
    os.makedirs(TEMP_UPLOAD_DIR, exist_ok=True)
    temp_path = os.path.join(TEMP_UPLOAD_DIR, f"upload_{file.filename}")
    
    try:
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)
        
        # Execute SQL
        from src.db import db
        await db.connect()
        rows = await db.fetch_all(sql)
        data = [dict(row) for row in rows]
        
        # Simple fill
        from openpyxl import load_workbook
        wb = load_workbook(temp_path)
        ws = wb.active
        
        # Get headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        
        # Find empty row
        data_row = ws.max_row + 1
        for row in range(2, ws.max_row + 100):
            if ws.cell(row=row, column=1).value is None:
                data_row = row
                break
        
        # Fill data
        for idx, row_data in enumerate(data):
            for col_idx, header in enumerate(headers, 1):
                if header and header in row_data:
                    ws.cell(row=data_row + idx, column=col_idx, value=row_data[header])
        
        # Save
        output_path = os.path.join(REPORT_OUTPUT_DIR, f"filled_{file.filename}")
        os.makedirs(REPORT_OUTPUT_DIR, exist_ok=True)
        wb.save(output_path)
        wb.close()
        
        # Upload
        result_url = await _upload_to_nextjs(output_path, file.filename)
        
        return {
            "success": True,
            "resultUrl": result_url,
            "rowCount": len(data),
            "message": f"Đã fill {len(data)} dòng"
        }
        
    except Exception as e:
        logger.error(f"❌ Quick fill error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)