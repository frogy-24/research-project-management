import time
from openpyxl import load_workbook
import pandas as pd
import requests
from pydantic import BaseModel
from typing import List, Optional
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
class DonViHanhChinh(BaseModel):
    ma: str
    ten: Optional[str] = None
    magoc: str
    malk: str
    truocsapnhap: str

class ThongTinDonViHanhChinh(BaseModel):
    id: int
    dientichkm2: Optional[str] = None
    dansonguoi: Optional[str] = None
    trungtamhc: Optional[str] = None
    truocsapnhap: Optional[str] = None
    con: Optional[str] = None
    ma: Optional[str] = None
    ten: Optional[str] = None
    magoc: Optional[str] = None
    malk: Optional[str] = None
    stat: int
    diachi: Optional[str] = None
    dthoai: Optional[str] = None
    cancu: Optional[str] = None
    tentinh: Optional[str] = None
    link: Optional[str] = None

def read_excel_data(file_path):
    try:
        df = pd.read_excel(file_path, dtype={'Mã số': str, 'Mã số thống kê': str})
        df = df.fillna('')
        return df

    except Exception as e:
        print(f"Có lỗi xảy ra: {e}")
        return None
    
def get_list_dvhc():
    response = requests.post("https://sapnhap.bando.com.vn/p.co_dvhc",data={"ma":0}, verify=False)
    response.raise_for_status()
    json_data = response.json()
    list_dvhc = [DonViHanhChinh(**item) for item in json_data]
    return list_dvhc

def get_info_dvhc(malk:str):
    response = requests.post("https://sapnhap.bando.com.vn/p.co_dvhc_id",data={"malk":malk},verify=False)
    response.raise_for_status()
    json_data = response.json()
    list_info = [ThongTinDonViHanhChinh(**item) for item in json_data]
    (info, )= list_info
    return info

def main():
    file_path = 'C:\\data.xlsx' 
    wb = load_workbook(file_path)
    ws = wb.active
    list_dvhc = get_list_dvhc()
    index = 0
    for row in ws.iter_rows(min_row=2, max_col=7):
        if row[5].value is not None and row[6].value is not None: continue
        ma_so_thong_ke = str(row[1].value)
        dvhc = next((x for x in list_dvhc if x.ma == ma_so_thong_ke), None)
        if dvhc is None:
            continue
        malk = dvhc.malk
        info_dvhc = get_info_dvhc(malk)
        dientich = info_dvhc.dientichkm2
        danso = info_dvhc.dansonguoi
        row[5].value = danso
        row[6].value = dientich
        wb.save(file_path)
        index += 1
        print(index)
    print("Success")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        quit()
